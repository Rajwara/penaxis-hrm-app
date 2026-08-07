import datetime as dt

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db
from ..security import hash_password
from ..deps import get_current_user, require_admin, require_super_admin
from ..storage import (
    save_upload,
    delete_upload,
    ALLOWED_IMAGE_EXT,
    ALLOWED_DOC_EXT,
    MAX_IMAGE_BYTES,
    MAX_DOC_BYTES,
)

router = APIRouter(prefix="/employees", tags=["employees"])


def _hidden_from(viewer: models.User, target: models.User) -> bool:
    """
    A super-admin account is invisible to everyone except themselves.
    Regular admins can't see, list, edit, or track them at all.
    """
    if not target.is_super_admin:
        return False
    return viewer.id != target.id


def _can_manage_cnic(user: models.User) -> bool:
    return bool(user.is_super_admin or user.can_view_cnic)


def _to_out_redacting_cnic(user: models.User, viewer: models.User) -> schemas.UserOut:
    """
    Build a UserOut for `user`, stripping CNIC fields unless the viewer has
    explicit CNIC access (super admin or granted can_view_cnic) or is
    looking at their own record. CNIC is sensitive — by default only the
    invisible super admin manages it; specific people can be granted access
    individually, but regular HR/Admin and managers don't get it by default.
    """
    data = schemas.UserOut.model_validate(user)
    if not _can_manage_cnic(viewer) and viewer.id != user.id:
        data.cnic_url = None
        data.cnic_original_name = None
    return data


@router.get("", response_model=list[schemas.UserOut])
def list_employees(
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    q = db.query(models.User).filter(models.User.is_active == 1)
    if not admin.is_super_admin:
        q = q.filter(
            (models.User.is_super_admin == False) | (models.User.id == admin.id)  # noqa: E712
        )
    users = q.order_by(models.User.id).all()
    for u in users:
        u.maybe_convert_to_permanent(db)

    return [_to_out_redacting_cnic(u, admin) for u in users]


@router.post("", response_model=schemas.UserOut, status_code=status.HTTP_201_CREATED)
def create_employee(
    payload: schemas.UserCreate,
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    existing = db.query(models.User).filter(models.User.email == payload.email).first()
    if existing:
        if existing.is_active:
            raise HTTPException(status_code=400, detail="Email already registered")
        # This email belongs to a previously removed (soft-deleted) account.
        # Email is unique at the DB level, so we can't create a second row
        # for it - reactivate this same record instead, which also
        # preserves their attendance/leave history under the same user_id.
        existing.name = payload.name
        existing.hashed_password = hash_password(payload.password)
        existing.role = payload.role
        existing.department = payload.department
        existing.position = payload.position
        existing.phone = payload.phone
        existing.leave_quota = payload.leave_quota
        existing.manager_id = payload.manager_id
        existing.employment_type = payload.employment_type
        existing.is_team_manager = payload.is_team_manager
        existing.is_active = 1
        db.commit()
        db.refresh(existing)
        return existing
    user = models.User(
        name=payload.name,
        email=payload.email,
        hashed_password=hash_password(payload.password),
        role=payload.role,
        department=payload.department,
        position=payload.position,
        phone=payload.phone,
        leave_quota=payload.leave_quota,
        manager_id=payload.manager_id,
        employment_type=payload.employment_type,
        is_team_manager=payload.is_team_manager,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.get("/my-team", response_model=list[schemas.UserOut])
def my_team(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Everyone who directly reports to the current user. Returns an empty
    list for anyone who isn't a manager — this isn't restricted to admins,
    it's just naturally empty if you don't manage anyone.
    """
    return [
        _to_out_redacting_cnic(u, current_user)
        for u in db.query(models.User)
        .filter(models.User.manager_id == current_user.id, models.User.is_active == 1)
        .order_by(models.User.name)
        .all()
    ]


@router.get("/{user_id}", response_model=schemas.UserOut)
def get_employee(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != models.Role.ADMIN and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized to view this profile")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if _hidden_from(current_user, user):
        raise HTTPException(status_code=404, detail="Employee not found")
    return _to_out_redacting_cnic(user, current_user)


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_employee(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="You cannot remove your own account")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if _hidden_from(admin, user):
        raise HTTPException(status_code=404, detail="Employee not found")
    # soft delete to preserve attendance/leave history
    user.is_active = 0
    db.commit()
    return None


@router.patch("/{user_id}/leave-quota", response_model=schemas.UserOut)
def update_leave_quota(
    user_id: int,
    payload: schemas.LeaveQuotaUpdate,
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    user.leave_quota = payload.leave_quota
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/reset-password", status_code=status.HTTP_204_NO_CONTENT)
def reset_password(
    user_id: int,
    payload: schemas.PasswordReset,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if _hidden_from(admin, user):
        raise HTTPException(status_code=404, detail="Employee not found")
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    user.hashed_password = hash_password(payload.new_password)
    db.commit()
    return None


@router.patch("/{user_id}", response_model=schemas.UserOut)
def update_employee(
    user_id: int,
    payload: schemas.UserUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != models.Role.ADMIN and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this profile")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if _hidden_from(current_user, user):
        raise HTTPException(status_code=404, detail="Employee not found")

    updates = payload.model_dump(exclude_unset=True)
    # Employees editing their own profile cannot change org-controlled fields
    if current_user.role != models.Role.ADMIN:
        updates.pop("department", None)
        updates.pop("position", None)
        updates.pop("manager_id", None)
        updates.pop("employment_type", None)
        updates.pop("permanent_conversion_date", None)
        updates.pop("is_team_manager", None)
        updates.pop("role", None)
        updates.pop("email", None)
        updates.pop("internship_end_date_override", None)
    elif "manager_id" in updates and updates["manager_id"] == user_id:
        raise HTTPException(status_code=400, detail="An employee cannot be their own manager")

    if "email" in updates and updates["email"] != user.email:
        existing = (
            db.query(models.User)
            .filter(models.User.email == updates["email"], models.User.id != user_id)
            .first()
        )
        if existing:
            raise HTTPException(status_code=400, detail="This email is already in use")

    for field, value in updates.items():
        setattr(user, field, value)  # skills uses the model's property setter
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/internship-feedback", response_model=schemas.UserOut)
def submit_internship_feedback(
    user_id: int,
    payload: schemas.InternshipFeedbackCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.id != user_id:
        raise HTTPException(
            status_code=403, detail="You can only submit your own internship feedback"
        )
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if user.employment_type != models.EmploymentType.INTERN:
        raise HTTPException(
            status_code=400, detail="This form only applies to internship accounts"
        )
    if not user.is_internship_completed:
        raise HTTPException(
            status_code=400, detail="Your internship period hasn't ended yet"
        )
    user.intern_feedback = payload.feedback
    user.intern_feedback_submitted_at = dt.datetime.utcnow()
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/profile-picture", response_model=schemas.UserOut)
async def upload_profile_picture(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != models.Role.ADMIN and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this profile")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    content = await file.read()
    if len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be under 5 MB")
    try:
        stored_name = save_upload(content, file.filename or "photo", ALLOWED_IMAGE_EXT)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    delete_upload(user.profile_picture)
    user.profile_picture = stored_name
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/cv", response_model=schemas.UserOut)
async def upload_cv(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != models.Role.ADMIN and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this profile")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    content = await file.read()
    if len(content) > MAX_DOC_BYTES:
        raise HTTPException(status_code=400, detail="File must be under 10 MB")
    try:
        stored_name = save_upload(content, file.filename or "cv", ALLOWED_DOC_EXT)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    delete_upload(user.cv_filename)
    user.cv_filename = stored_name
    user.cv_original_name = file.filename
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/cnic", response_model=schemas.UserOut)
async def upload_cnic(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    # CNIC is sensitive. Employees can submit their own once; after that,
    # only someone with explicit CNIC access (super admin, or specifically
    # granted) can replace or remove it — regular HR/Admin staff don't have
    # access by default, even though they can manage everything else.
    if current_user.id != user_id and not _can_manage_cnic(current_user):
        raise HTTPException(status_code=403, detail="Not authorized to edit this profile")
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    if current_user.id == user_id and not _can_manage_cnic(current_user) and user.cnic_filename:
        raise HTTPException(
            status_code=400,
            detail="Your CNIC has already been submitted. Contact HR if it needs to be corrected.",
        )

    content = await file.read()
    if len(content) > MAX_DOC_BYTES:
        raise HTTPException(status_code=400, detail="File must be under 10 MB")
    try:
        stored_name = save_upload(content, file.filename or "cnic", ALLOWED_IMAGE_EXT | ALLOWED_DOC_EXT)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    delete_upload(user.cnic_filename)
    user.cnic_filename = stored_name
    user.cnic_original_name = file.filename
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/promote-super-admin", response_model=schemas.UserOut)
def promote_super_admin(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_super_admin),
):
    """
    Grants full admin rights plus total invisibility to every other admin/HR
    user. This is deliberately a separate, explicit action rather than a
    field on the general edit form — it's too sensitive to be a checkbox
    anyone with HR access could casually flip. Only an existing super admin
    can grant this to someone else; a regular Admin/HR account cannot.
    """
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    user.role = models.Role.ADMIN
    user.is_super_admin = True
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/demote-super-admin", response_model=schemas.UserOut)
def demote_super_admin(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_super_admin),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    user.is_super_admin = False
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/grant-cnic-access", response_model=schemas.UserOut)
def grant_cnic_access(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_super_admin),
):
    """
    Grants a specific Admin/HR user the ability to view, upload, and replace
    CNIC documents for anyone. Deliberately separate from the Admin role
    itself and from super-admin status - only the super admin can grant or
    revoke this, on a per-person basis, so it stays limited to whoever is
    explicitly trusted with it rather than every HR account by default.
    """
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    user.can_view_cnic = True
    db.commit()
    db.refresh(user)
    return user


@router.post("/{user_id}/revoke-cnic-access", response_model=schemas.UserOut)
def revoke_cnic_access(
    user_id: int,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_super_admin),
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    user.can_view_cnic = False
    db.commit()
    db.refresh(user)
    return user


def _generate_password(length: int = 10) -> str:
    import secrets
    import string

    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _infer_department(title: str) -> str:
    import re

    t = title.lower()

    def has(*keywords: str) -> bool:
        return any(re.search(r"\b" + re.escape(k) + r"\b", t) for k in keywords)

    if has("ceo", "founder", "co-founder", "coo"):
        return "Leadership"
    if has("marketing", "social media", "creative", "brand"):
        return "Marketing"
    if has("sales", "business development", "growth", "bd"):
        return "Sales & Growth"
    if has("developer", "engineer", "cms", "salesforce", "crm", "data", "hrm"):
        return "Engineering"
    if "human resource" in t:
        return "HR"
    if has("project manager"):
        return "Operations"
    if has("client success"):
        return "Customer Success"
    if has("business analyst", "business operations", "accounts"):
        return "Operations"
    return "General"


@router.post("/bulk-import")
async def bulk_import_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """
    Creates accounts in bulk from an uploaded roster spreadsheet. Expects a
    header row containing (at least) Name and Title/Position columns; Email,
    Joining Date columns are used when present. Any row with the literal
    text "intern" anywhere in it is treated as an internship. Rows whose
    joining-date cell reads like "not an onsite resource" are treated as
    contractors with no fixed join date (defaults to today).

    Generates a random password per new account and returns it in plaintext
    exactly once, in this response — there is no way to retrieve it again
    afterward, so save/share it immediately.
    """
    from openpyxl import load_workbook
    import io

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read this file as an Excel spreadsheet")
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="This spreadsheet appears to be empty")

    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    def find_col(*keywords: str) -> int | None:
        for i, h in enumerate(header):
            if any(k in h for k in keywords):
                return i
        return None

    name_i = find_col("name")
    title_i = find_col("title")
    join_i = find_col("joining")
    email_i = find_col("email")

    if name_i is None:
        raise HTTPException(status_code=400, detail="Couldn't find a 'Name' column in this file")

    results = []
    for row in rows[1:]:
        if not row or name_i >= len(row) or not row[name_i]:
            continue
        name = str(row[name_i]).strip()
        title = (
            str(row[title_i]).strip()
            if title_i is not None and title_i < len(row) and row[title_i]
            else "Staff"
        )
        raw_email = (
            str(row[email_i]).strip()
            if email_i is not None and email_i < len(row) and row[email_i]
            else None
        )
        email_guessed = False
        if raw_email:
            email = raw_email
        else:
            email = f"{name.lower().replace(' ', '')}penaxis@gmail.com"
            email_guessed = True

        is_intern = any(
            isinstance(c, str) and c.strip().lower() == "intern" for c in row
        )
        raw_join = row[join_i] if join_i is not None and join_i < len(row) else None
        is_contract = isinstance(raw_join, str) and "not an onsite" in raw_join.lower()

        if isinstance(raw_join, dt.datetime):
            join_date = raw_join.date()
        elif isinstance(raw_join, dt.date):
            join_date = raw_join
        else:
            join_date = dt.date.today()

        employment_type = (
            models.EmploymentType.INTERN
            if is_intern
            else models.EmploymentType.CONTRACT
            if is_contract
            else models.EmploymentType.PERMANENT
        )

        existing = db.query(models.User).filter(models.User.email == email).first()
        if existing:
            results.append({
                "name": name,
                "email": email,
                "password": None,
                "status": "skipped",
                "note": "An account with this email already exists",
            })
            continue

        password = _generate_password()
        user = models.User(
            name=name,
            email=email,
            hashed_password=hash_password(password),
            role=models.Role.EMPLOYEE,
            department=_infer_department(title),
            position=title,
            join_date=join_date,
            employment_type=employment_type,
        )
        db.add(user)
        results.append({
            "name": name,
            "email": email,
            "password": password,
            "status": "created",
            "note": "Email was guessed (not listed in the file) — please verify it" if email_guessed else "",
        })

    db.commit()
    return {"results": results}
