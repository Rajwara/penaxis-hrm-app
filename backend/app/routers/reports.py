import io
import datetime as dt
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .. import models
from ..database import get_db
from ..deps import require_admin

router = APIRouter(prefix="/reports", tags=["reports"])

# The app stores/works in naive UTC internally; the whole org is on this timezone.
DISPLAY_TZ = ZoneInfo("Asia/Karachi")


def _to_local(value: dt.datetime | None) -> dt.datetime | None:
    """Naive UTC -> aware Asia/Karachi, for display in the exported file."""
    if value is None:
        return None
    return value.replace(tzinfo=ZoneInfo("UTC")).astimezone(DISPLAY_TZ)

HEADER_FILL = PatternFill(start_color="734FA0", end_color="734FA0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="212121")


def _style_header(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, ncols: int, min_width: int = 12, max_width: int = 40):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_len, max_width)


def _hours_worked(check_in, check_out) -> str:
    if not check_in or not check_out:
        return ""
    delta = check_out - check_in
    return f"{delta.total_seconds() / 3600:.1f}"


@router.get("/employees-excel")
def export_employees_excel(
    start_date: dt.date | None = Query(None, description="Filter attendance/leaves from this date (inclusive)"),
    end_date: dt.date | None = Query(None, description="Filter attendance/leaves up to this date (inclusive)"),
    employee_id: int | None = Query(None, description="Limit the report to a single employee"),
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    if start_date and end_date and end_date < start_date:
        start_date, end_date = end_date, start_date

    employees_q = db.query(models.User).filter(models.User.is_active == 1)
    if employee_id:
        employees_q = employees_q.filter(models.User.id == employee_id)
        if employees_q.count() == 0:
            raise HTTPException(status_code=404, detail="Employee not found")
    employees = employees_q.order_by(models.User.department, models.User.name).all()

    wb = Workbook()

    # --- Sheet 1: Employees overview ---
    ws1 = wb.active
    ws1.title = "Employees"
    headers1 = [
        "Name", "Email", "Department", "Position", "Role",
        "Phone", "Join Date", "Leave Quota (days)",
    ]
    ws1.append(headers1)
    _style_header(ws1, 1, len(headers1))
    for emp in employees:
        ws1.append([
            emp.name, emp.email, emp.department, emp.position,
            emp.role.value, emp.phone or "", emp.join_date.isoformat(),
            emp.leave_quota,
        ])
    ws1.freeze_panes = "A2"
    _autofit(ws1, len(headers1))

    # --- Sheet 2: Attendance ---
    ws2 = wb.create_sheet("Attendance")
    headers2 = [
        "Employee", "Department", "Date", "Check In (PKT)", "Check Out (PKT)", "Hours Worked",
    ]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    attendance_q = (
        db.query(models.Attendance)
        .join(models.User)
        .filter(models.User.is_active == 1)
    )
    if employee_id:
        attendance_q = attendance_q.filter(models.Attendance.user_id == employee_id)
    if start_date:
        attendance_q = attendance_q.filter(models.Attendance.date >= start_date)
    if end_date:
        attendance_q = attendance_q.filter(models.Attendance.date <= end_date)
    attendance_rows = attendance_q.order_by(
        models.Attendance.date.desc(), models.Attendance.check_in.desc()
    ).all()
    for rec in attendance_rows:
        check_in_local = _to_local(rec.check_in)
        check_out_local = _to_local(rec.check_out)
        ws2.append([
            rec.user.name,
            rec.user.department,
            rec.date.isoformat(),
            check_in_local.strftime("%Y-%m-%d %H:%M") if check_in_local else "",
            check_out_local.strftime("%Y-%m-%d %H:%M") if check_out_local else "",
            _hours_worked(rec.check_in, rec.check_out),
        ])
    ws2.freeze_panes = "A2"
    _autofit(ws2, len(headers2))

    # --- Sheet 3: Leaves ---
    ws3 = wb.create_sheet("Leaves")
    headers3 = [
        "Employee", "Department", "Leave Type", "Start Date", "End Date",
        "Days", "Status", "Reason", "Submitted At (PKT)", "Decided At (PKT)",
    ]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    leave_q = (
        db.query(models.LeaveRequest)
        .join(models.User)
        .filter(models.User.is_active == 1)
    )
    if employee_id:
        leave_q = leave_q.filter(models.LeaveRequest.user_id == employee_id)
    if start_date:
        # include leaves that overlap the range at all, not just ones starting inside it
        leave_q = leave_q.filter(models.LeaveRequest.end_date >= start_date)
    if end_date:
        leave_q = leave_q.filter(models.LeaveRequest.start_date <= end_date)
    leave_rows = leave_q.order_by(models.LeaveRequest.created_at.desc()).all()
    for lv in leave_rows:
        submitted_local = _to_local(lv.created_at)
        decided_local = _to_local(lv.decided_at)
        ws3.append([
            lv.user.name,
            lv.user.department,
            lv.leave_type.value,
            lv.start_date.isoformat(),
            lv.end_date.isoformat(),
            lv.days,
            lv.status.value,
            lv.reason,
            submitted_local.strftime("%Y-%m-%d %H:%M") if submitted_local else "",
            decided_local.strftime("%Y-%m-%d %H:%M") if decided_local else "",
        ])
    ws3.freeze_panes = "A2"
    _autofit(ws3, len(headers3))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if start_date and end_date:
        suffix = f"{start_date.isoformat()}_to_{end_date.isoformat()}"
    else:
        suffix = dt.date.today().isoformat()
    if employee_id and employees:
        name_slug = employees[0].name.lower().replace(" ", "-")
        suffix = f"{name_slug}-{suffix}"
    filename = f"penaxis-hr-report-{suffix}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
